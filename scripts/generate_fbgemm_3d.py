import math
import openpyxl
wb=openpyxl.load_workbook("/home/huiwu1/workspace/int8/mkl-dnn/test_file_3d.xlsx")
f=open(r"shape_conv3d_fbgemm","w")
#wb=openpyxl.load_workbook("/home/huiwu1/workspace/int8/mkl-dnn/test_file_dw.xlsx")
#f=open(r"shape_conv_dw","a+")

ws=wb.active
max_row = ws.max_row
max_col = ws.max_column 
for i in range(1,max_row+1):
    #for j in range(max_col):
    mb = ws.cell(row=i, column=1).value
    ic = ws.cell(row=i, column=2).value
    oc = ws.cell(row=i, column=3).value
    it = ws.cell(row=i, column=4).value
    ih = ws.cell(row=i, column=5).value
    iw = ws.cell(row=i, column=6).value

    g = ws.cell(row=i, column=7).value

    kt = ws.cell(row=i, column=8).value
    kh = ws.cell(row=i, column=9).value
    kw = ws.cell(row=i, column=10).value

    stride_t = ws.cell(row=i, column=11).value
    stride_h = ws.cell(row=i, column=12).value
    stride_w = ws.cell(row=i, column=13).value
    
    pad_prev = ws.cell(row=i, column=14).value
    pad_h_top = ws.cell(row=i, column=15).value
    pad_w_left = ws.cell(row=i, column=16).value
    pad_next = ws.cell(row=i, column=17).value
    pad_h_bottom = ws.cell(row=i, column=18).value
    pad_w_right = ws.cell(row=i, column=19).value

    d_t = ws.cell(row=i, column=20).value
    d_h = ws.cell(row=i, column=21).value
    d_w = ws.cell(row=i, column=22).value
 
    if not ih:
        continue

    if not pad_prev:
        pad_prev = 0
    if not pad_h_top:
        pad_h_top = 0
    if not pad_w_left:
        pad_w_left = 0
    if not pad_next:
        pad_next = 0
    if not pad_h_bottom:
        pad_h_bottom = 0
    if not pad_w_right:
        pad_w_right = 0
    if not d_t:
        d_t = 1
    if not d_h:
        d_h = 1
    if not d_w:
        d_w = 1
    newline = "conv_param_t<3>(" + \
        str(mb) + "," + str(ic) + "," + str(oc) + "," + \
        "{" + str(it) + "," + str(ih) + "," + str(iw) + "}" + "," + str(g) + "," + \
        "{" + str(kt) + "," + str(kh) + "," + str(kw) + "}" + "," + \
        "{" + str(stride_t) + "," + str(stride_h) + "," + str(stride_w) + "}," + \
        "{" + str(pad_prev) + "," + str(pad_h_top) + "," + str(pad_w_left) + "," + \
              str(pad_next) + "," + str(pad_h_bottom) + "," + str(pad_w_right) + "}," + \
        "{" + str(d_t) + "," + str(d_h) + "," + str(d_w) + "}),";

    f.writelines(newline+"\n")

f.close()
